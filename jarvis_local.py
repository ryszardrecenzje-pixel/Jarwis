import ollama
import openpyxl
import pyautogui
import time
import os
import json
import csv
import pyperclip
from pathlib import Path
from openpyxl import Workbook, load_workbook
from datetime import datetime

# ================== KONFIGURACJA ==================
MODEL = "qwen2.5:7b"              # zmień na swój model jeśli chcesz
DOWNLOADS_PATH = str(Path.home() / "Downloads")  # folder Pobrane
pyautogui.FAILSAFE = True         # mysz w lewy górny róg = STOP
pyautogui.PAUSE = 0.25
# ==================================================

SYSTEM_PROMPT = """
Jesteś lokalnym asystentem automatyzacji na Windows.
Odpowiadasz WYŁĄCZNIE poprawnym JSON-em:

{
  "actions": [
    {"type": "nazwa_akcji", ...},
    ...
  ]
}

Dostępne akcje:

1. read_file
   {"type": "read_file", "path": "pełna_ścieżka"}

2. create_excel
   {"type": "create_excel", "path": "ścieżka.xlsx", "data": [["A1", "B1"], ["A2", "B2"]]}

3. edit_excel
   {"type": "edit_excel", "path": "istniejący.xlsx", "cell": "A1", "value": "nowa wartość"}
   lub
   {"type": "edit_excel", "path": "istniejący.xlsx", "data": [["A1", "B1"], ["wartość1", "wartość2"]], "start_row": 1}

4. type_text
   {"type": "type_text", "text": "tekst"}

5. click
   {"type": "click", "x": 100, "y": 200}

6. hotkey
   {"type": "hotkey", "keys": ["ctrl", "c"]}

7. open_app
   {"type": "open_app", "name": "excel"}  
   Możliwe: excel, notepad, explorer, chrome, calc, word, edge

8. wait
   {"type": "wait", "seconds": 1.5}

9. copy_to_clipboard
   {"type": "copy_to_clipboard", "text": "tekst do skopiowania"}

10. get_newest_file
    {"type": "get_newest_file", "folder": "C:\\\\Users\\\\...\\\\Downloads", "extension": ".csv"}
    (extension jest opcjonalne)

11. done
    {"type": "done", "message": "Zadanie wykonane"}

Zasady:
- Zawsze zwracaj TYLKO czysty JSON.
- Ścieżki Windows zapisuj z podwójnymi backslashami: C:\\\\Users\\\\Nazwa\\\\Desktop\\\\plik.xlsx
- Jeśli nie wiesz jak coś zrobić – użyj done z komunikatem.
"""

def execute_action(action: dict) -> str:
    t = action.get("type")

    try:
        # ---------- CZYTANIE PLIKÓW ----------
        if t == "read_file":
            path = action["path"]
            if not os.path.exists(path):
                return f"Błąd: plik nie istnieje → {path}"

            # CSV
            if path.lower().endswith(".csv"):
                with open(path, "r", encoding="utf-8", errors="ignore") as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                return f"CSV wczytany ({len(rows)} wierszy). Pierwsze wiersze:\n{rows[:8]}"

            # Excel .xlsx
            if path.lower().endswith((".xlsx", ".xlsm")):
                wb = load_workbook(path, data_only=True)
                ws = wb.active
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(list(row))
                return f"Excel wczytany ({len(data)} wierszy). Pierwsze wiersze:\n{data[:8]}"

            # Zwykły tekst
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                content = f.read()
            return f"Plik tekstowy wczytany ({len(content)} znaków):\n{content[:600]}..."

        # ---------- TWORZENIE EXCELA ----------
        elif t == "create_excel":
            path = action["path"]
            data = action.get("data", [])

            wb = Workbook()
            ws = wb.active
            ws.title = "Dane"

            for row in data:
                ws.append(row)

            Path(path).parent.mkdir(parents=True, exist_ok=True)
            wb.save(path)
            return f"Utworzono nowy Excel: {path}"

        # ---------- EDYCJA EXCELA ----------
        elif t == "edit_excel":
            path = action["path"]
            if not os.path.exists(path):
                return f"Błąd: plik nie istnieje → {path}"

            wb = load_workbook(path)
            ws = wb.active

            # Pojedyncza komórka
            if "cell" in action and "value" in action:
                ws[action["cell"]] = action["value"]
                wb.save(path)
                return f"Zmieniono komórkę {action['cell']} w pliku {path}"

            # Wiele wierszy
            if "data" in action:
                start_row = action.get("start_row", 1)
                for i, row in enumerate(action["data"]):
                    for j, value in enumerate(row):
                        ws.cell(row=start_row + i, column=j + 1, value=value)
                wb.save(path)
                return f"Zaktualizowano dane w pliku {path}"

            return "Brak danych do edycji"

        # ---------- PISANIE ----------
        elif t == "type_text":
            text = action.get("text", "")
            pyautogui.write(text, interval=0.02)
            return f"Wpisano tekst ({len(text)} znaków)"

        # ---------- KLIKANIE ----------
        elif t == "click":
            x = int(action["x"])
            y = int(action["y"])
            pyautogui.click(x, y)
            return f"Kliknięto w ({x}, {y})"

        # ---------- SKRÓTY ----------
        elif t == "hotkey":
            keys = action.get("keys", [])
            pyautogui.hotkey(*keys)
            return f"Wysłano skrót: {' + '.join(keys)}"

        # ---------- OTWIERANIE APLIKACJI ----------
        elif t == "open_app":
            name = action.get("name", "").lower()
            apps = {
                "excel": "excel",
                "notepad": "notepad",
                "explorer": "explorer",
                "chrome": "chrome",
                "edge": "msedge",
                "calc": "calc",
                "word": "winword",
            }
            if name in apps:
                os.system(f"start {apps[name]}")
                time.sleep(1.8)
                return f"Otwarto: {name}"
            return f"Nie znam aplikacji: {name}"

        # ---------- CZEKANIE ----------
        elif t == "wait":
            seconds = float(action.get("seconds", 1))
            time.sleep(seconds)
            return f"Czekałem {seconds}s"

        # ---------- SCHOWEK ----------
        elif t == "copy_to_clipboard":
            text = action.get("text", "")
            pyperclip.copy(text)
            return "Tekst skopiowany do schowka"

        # ---------- NAJNOWSZY PLIK ----------
        elif t == "get_newest_file":
            folder = action.get("folder", DOWNLOADS_PATH)
            extension = action.get("extension", None)

            if not os.path.exists(folder):
                return f"Folder nie istnieje: {folder}"

            files = []
            for f in os.listdir(folder):
                full = os.path.join(folder, f)
                if os.path.isfile(full):
                    if extension is None or f.lower().endswith(extension.lower()):
                        files.append(full)

            if not files:
                return "Nie znaleziono żadnych plików"

            newest = max(files, key=os.path.getmtime)
            return f"Najnowszy plik: {newest}"

        # ---------- KONIEC ----------
        elif t == "done":
            return action.get("message", "Gotowe")

        else:
            return f"Nieznana akcja: {t}"

    except Exception as e:
        return f"Błąd przy akcji '{t}': {str(e)}"


def main():
    print("=" * 65)
    print("  Lokalny Jarvis – Windows + Ollama")
    print("  Komendy tekstowe | Pełna kontrola plików i Excela")
    print("  Wpisz 'exit' żeby wyjść")
    print("  FAILSAFE: przesuń myszkę w lewy górny róg = natychmiastowy STOP")
    print("=" * 65)

    while True:
        user_input = input("\nTy > ").strip()
        if user_input.lower() in ["exit", "quit", "wyjście", "koniec"]:
            print("Do zobaczenia!")
            break

        if not user_input:
            continue

        print("Myślę...")

        try:
            response = ollama.chat(
                model=MODEL,
                messages=[
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {"role": "user", "content": user_input}
                ],
                format="json",
                options={"temperature": 0.15}
            )

            content = response["message"]["content"]
            print("\n--- Odpowiedź modelu ---")
            print(content)
            print("------------------------")

            data = json.loads(content)
            actions = data.get("actions", [])

            if not actions:
                print("Brak akcji do wykonania.")
                continue

            print("\nWykonuję:")
            for i, action in enumerate(actions, 1):
                result = execute_action(action)
                print(f"  {i}. {result}")

        except json.JSONDecodeError:
            print("Model nie zwrócił poprawnego JSON-a. Spróbuj inaczej sformułować komendę.")
        except Exception as e:
            print(f"Błąd ogólny: {e}")


if __name__ == "__main__":
    main()
